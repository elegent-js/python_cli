#!/Users/liupeijun/Documents/40_Cli/python_cli/venv/bin/python3

import mysql.connector
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.styles import PatternFill
import os
import utils.util as util


def fetch_tables(cursor):
    """
    Fetch the list of tables and their comments from the database.
    """
    tables_query = """
        SELECT 
            table_name, 
            table_comment 
        FROM information_schema.tables 
        WHERE table_schema = 'build_in_dgmp';
    """
    cursor.execute(tables_query)
    return cursor.fetchall()

def fetch_columns(cursor, table_name):
    """
    Fetch the columns and their details for a given table.
    """
    query = f"""
    SELECT
        extra,
        column_name,
        column_type,
        is_nullable,
        column_default,
        column_comment
    FROM
        information_schema.COLUMNS 
    WHERE
        table_name = '{table_name}'
        AND table_schema = 'build_in_dgmp'
        ORDER BY ordinal_position
    """
    cursor.execute(query)
    return cursor.fetchall()

def load_template(template_path):
    """
    Load the Excel template from the given path.
    """
    return openpyxl.load_workbook(template_path)

def copy_template_sheet(wb, template_sheet, table_name, table_comment):
    """
    Copy the template sheet and rename it based on the table name or comment.
    """
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = table_comment if table_comment else table_name
    return new_sheet

def fill_sheet(new_sheet, columns, style_row, fill):
    """
    Fill the new sheet with column data and apply styles.
    """
    row = 5
    for column in columns:
        # Copy styles from the template row
        for col_num, cell in enumerate(style_row, 1):
            new_cell = new_sheet.cell(row=row, column=col_num)
            new_cell._style = copy(cell._style)

        # Apply fill to specific columns
        if column[1] in ['create_time', 'update_time', 'delete_time', 'delete_flag', 'create_user_name', 'update_user_name', 'delete_user_name']:
            for col_num, cell in enumerate(style_row, 1):
                if col_num > 1:
                    new_cell = new_sheet.cell(row=row, column=col_num)
                    new_cell.fill = fill

        # Fill the sheet with column data
        new_sheet[f'B{row}'] = '●' if (column[0] == 'auto_increment' or column[5].upper() == 'PK') else ''
        new_sheet[f'C{row}'] = column[1]
        new_sheet[f'D{row}'] = column[2]
        new_sheet[f'E{row}'] = '○' if column[3] == 'NO' else ''
        new_sheet[f'F{row}'] = column[4]
        new_sheet[f'G{row}'] = column[5]

        row += 1

def adjust_column_width(new_sheet):
    """
    Adjust the width of the columns in the new sheet based on the content.
    """
    for col in new_sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 6)
        new_sheet.column_dimensions[column].width = adjusted_width

def main():
    """
    Main function to generate the Excel file with database structure.
    """
    args = util.getArgs()

    connection = util.connect_to_database(args.host, args.user, args.password, args.database, args.port)
    cursor = connection.cursor()

    tables = fetch_tables(cursor)
    # Use relative path for the template
    template_path = os.path.join(os.path.dirname(__file__), 'python_cli/export_to_excel/files', 'template.xlsx')
    wb = load_template(template_path)
    
    template_sheet = wb.active
    style_row = template_sheet[5]
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for table in tables:
        table_name, table_comment = table
        new_sheet = copy_template_sheet(wb, template_sheet, table_name, table_comment)
        columns = fetch_columns(cursor, table_name)
        new_sheet['C2'] = table_name
        new_sheet['D2'] = table_comment
        fill_sheet(new_sheet, columns, style_row, fill)
        adjust_column_width(new_sheet)

    wb.remove(template_sheet)
    wb.save(args.output)
    connection.close()



if __name__ == "__main__":
    main()